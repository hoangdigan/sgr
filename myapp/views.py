import os
import mimetypes
import pandas as pd
from django.http.response import HttpResponse
from django.shortcuts import render
from datetime import datetime

import mysql.connector
from mysql.connector import Error
import openpyxl
import openpyxl as xl
import openpyxl.utils.cell 
from openpyxl import Workbook

def index(request):    
    global stock 
    if request.method == 'POST':
        stock = request.POST['stock']        
        dataprocess(stock)
    return render(request, "index.html")      

def institutions(request):
    global stock 
    if request.method == 'POST':
        stock = request.POST['stock']        
        dataprocess(stock)
    return render(request, "institutions.html") 

def bond(request):  
    global stock   
    stock = "bond"    
    return render(request, "bond.html") 

def industry(request):
    global stock   
    stock = "industrial"  
    return render(request, "industrial.html") 

def macro(request):  
    global stock   
    stock = "sbvex"
    return render(request, "macro.html") 

def benchmark(request):
    global stock   
    stock = "benchmark"
    return render(request, "benchmark.html") 

def other(request):
    global stock 
    stock = "badd"
    return render(request, "other.html") 

def dataprocess(stock):
    try:
        conn = mysql.connector.connect(host='localhost',
                            database='saigonrates',
                            user='root',
                            password='123456')

        cursor = conn.cursor()

        sql =   f"""select * from financial_statement WHERE financial_statement_id like '{stock}%'""" 
        cursor.execute(sql)       
        df = pd.DataFrame.from_records(cursor.fetchall(),
                                columns = [desc[0] for desc in cursor.description])       

        # Define Django project base directory    
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))

        # Define text file name
        filename = f'{stock}_proccess.xlsx'

        # Define the full file path
        filepath = os.path.join(BASE_DIR,'download',filename)   

        # company profile

        sql =   f"""select * from company_profile WHERE stock_code like '{stock}'""" 
        cursor.execute(sql)       
        records = cursor.fetchall() 

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        wb.create_sheet('company_profile')
        ws = wb['company_profile']

        for rc in records:
            ws.cell(1,1).value = "MST"
            ws.cell(1,2).value = rc[0]

            ws.cell(2,1).value = "Company Name"
            ws.cell(2,2).value = rc[1]

            ws.cell(3,1).value = "English Name"
            ws.cell(3,2).value = rc[2]

            ws.cell(4,1).value = "MACK"
            ws.cell(4,2).value = rc[3]

            ws.cell(5,1).value = "Declaration Person"
            ws.cell(5,2).value = rc[4]

            ws.cell(6,1).value = "Stage"
            ws.cell(6,2).value = rc[5]

            ws.cell(7,1).value = "Industry Group"
            ws.cell(7,2).value = rc[7]

            ws.cell(8,1).value = "Industry"
            ws.cell(8,2).value = rc[8]

            ws.cell(9,1).value = "Address"
            ws.cell(9,2).value = rc[12]

            ws.cell(10,1).value = "Phone"
            ws.cell(10,2).value = rc[13]

            ws.cell(11,1).value = "Fax"
            ws.cell(11,2).value = rc[14]

            ws.cell(12,1).value = "Email"
            ws.cell(12,2).value = rc[15]

            ws.cell(13,1).value = "Website"
            ws.cell(13,2).value = rc[16]

            ws.cell(14,1).value = "Capital"
            ws.cell(14,2).value = rc[17]

        wb.save(filepath)   
        
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:  
            df.to_excel(writer, sheet_name='Original FS')

    except Error as e:
        print(e)

    finally:
        cursor.close()
        conn.close()
        pass

def download_file(request):
    
    # Define Django project base directory    
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    # Define text file name
    filename = f'{stock}_proccess.xlsx'

    # Define the full file path
    filepath = os.path.join(BASE_DIR,'download',filename)

    # Open the file for reading content
    path = open(filepath, "rb")

    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)

    # Set the return value of the HttpResponse
    # response = HttpResponse(path, content_type=mime_type)
    response = HttpResponse(path, content_type="application/ms-excel")

    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # Return the response value
    return response


